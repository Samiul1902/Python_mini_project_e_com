import openpyxl

# Product Class
class Product:
    def __init__(self, product_id, name, price, stock):
        self.product_id = product_id
        self.name = name
        self.price = price
        self.stock = stock

    def update_stock(self, quantity):
        self.stock += quantity

    def to_dict(self):
        return {
            "Product ID": self.product_id,
            "Name": self.name,
            "Price": self.price,
            "Stock": self.stock
        }

# Customer Class
class Customer:
    def __init__(self, customer_id, name, email):
        self.customer_id = customer_id
        self.name = name
        self.email = email

    def to_dict(self):
        return {
            "Customer ID": self.customer_id,
            "Name": self.name,
            "Email": self.email
        }

# PremiumCustomer Class
class PremiumCustomer(Customer):
    def apply_discount(self, total):
        return total * 0.9  # 10% discount

# Order Class
class Order:
    def __init__(self, order_id, customer, products):
        self.order_id = order_id
        self.customer = customer
        self.products = products
        self.total_amount = self.calculate_total()

    def calculate_total(self):
        return sum(product["price"] * product["quantity"] for product in self.products)

    def to_dict(self):
        return {
            "Order ID": self.order_id,
            "Customer": self.customer.name,
            "Products": [product["name"] for product in self.products],
            "Total Amount": self.total_amount
        }

# Functions for Excel Handling
def load_data(filename, sheet_name):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        headers = rows[0]
        data = [dict(zip(headers, row)) for row in rows[1:]]
        return data
    except FileNotFoundError:
        return []

def save_data(filename, sheet_name, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    if data:
        headers = data[0].keys()
        sheet.append(list(headers))
        for record in data:
            sheet.append(list(record.values()))
    
    workbook.save(filename)

# Main Program Logic
products = [Product(**p) for p in load_data("products.xlsx", "Products")]
customers = [Customer(**c) for c in load_data("customers.xlsx", "Customers")]
orders = []

def main_menu():
    while True:
        print("\nWelcome to the E-Commerce System!")
        print("1. Add a new product")
        print("2. Register a new customer")
        print("3. Place an order")
        print("4. View order details")
        print("5. Exit")
        choice = input("Choose an option: ")

        if choice == "1":
            add_product()
        elif choice == "2":
            register_customer()
        elif choice == "3":
            place_order()
        elif choice == "4":
            view_orders()
        elif choice == "5":
            exit_program()
            break
        else:
            print("Invalid choice. Please try again.")

def add_product():
    product_id = input("Enter Product ID: ")
    name = input("Enter Product Name: ")
    price = float(input("Enter Product Price: "))
    stock = int(input("Enter Product Stock: "))
    products.append(Product(product_id, name, price, stock))
    print("Product added successfully!")

def register_customer():
    customer_id = input("Enter Customer ID: ")
    name = input("Enter Customer Name: ")
    email = input("Enter Customer Email: ")
    customers.append(Customer(customer_id, name, email))
    print("Customer registered successfully!")

def place_order():
    customer_id = input("Enter Customer ID: ")
    customer = next((c for c in customers if c.customer_id == customer_id), None)

    if not customer:
        print("Customer not found!")
        return

    order_id = f"O{len(orders) + 1:03}"
    products_ordered = []

    while True:
        product_id = input("Enter Product ID (or 'done' to finish): ")
        if product_id.lower() == "done":
            break
        product = next((p for p in products if p.product_id == product_id), None)
        
        if not product:
            print("Product not found!")
            continue

        quantity = int(input("Enter Quantity: "))
        if quantity > product.stock:
            print("Not enough stock available!")
            continue

        product.stock -= quantity
        products_ordered.append({"name": product.name, "price": product.price, "quantity": quantity})

    if products_ordered:
        total_amount = sum(p["price"] * p["quantity"] for p in products_ordered)
        if isinstance(customer, PremiumCustomer):
            total_amount = customer.apply_discount(total_amount)

        orders.append(Order(order_id, customer, products_ordered))
        print(f"Order placed successfully! Total Amount: ${total_amount:.2f}")

def view_orders():
    if not orders:
        print("No orders found!")
        return

    for order in orders:
        print("\nOrder Details:")
        print(f"Order ID: {order.order_id}")
        print(f"Customer: {order.customer.name}")
        print(f"Products: {', '.join(p['name'] for p in order.products)}")
        print(f"Total Amount: ${order.total_amount:.2f}")

def exit_program():
    save_data("products.xlsx", "Products", [p.to_dict() for p in products])
    save_data("customers.xlsx", "Customers", [c.to_dict() for c in customers])
    print("Data saved successfully. Exiting... Goodbye!")

# Run the program
if __name__ == "__main__":
    main_menu()
